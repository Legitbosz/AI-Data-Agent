"""
validator.py — Data Validation & Quality Engine for AI Data Analysis Agent

13 functions:
 1. Required field validation
 2. Email validation
 3. Phone validation
 4. Duplicate detection
 5. Error detection
 6. Error notes
 7. Overall status
 8. Workflow status
 9. Auto completion date
10. Conditional formatting (returns rules for UI/export)
11. Error rate calculation
12. Data accuracy calculation
13. Productivity tracker
"""

import pandas as pd
import numpy as np
import re
from datetime import datetime

# ─── PATTERN DEFINITIONS ─────────────────────────────────────────────────────

EMAIL_REGEX = re.compile(
    r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$"
)

PHONE_REGEX = re.compile(
    r"^[\+]?[\d\s\-\(\)\.]{7,20}$"
)

# Column name patterns for auto-detection
EMAIL_COL_PATTERNS = re.compile(
    r"e[\-_\s]?mail|email[\-_\s]?addr", re.IGNORECASE
)
PHONE_COL_PATTERNS = re.compile(
    r"phone|mobile|cell|tel(?:ephone)?|contact[\-_\s]?num|gsm|whatsapp", re.IGNORECASE
)
DATE_COL_PATTERNS = re.compile(
    r"date|_dt$|_at$|timestamp|created|updated|completed|deadline|due|start|end|time",
    re.IGNORECASE,
)
STATUS_COL_PATTERNS = re.compile(
    r"status|state|stage|progress|phase|workflow|step", re.IGNORECASE
)
NAME_COL_PATTERNS = re.compile(
    r"name|first[\-_\s]?name|last[\-_\s]?name|full[\-_\s]?name|customer|client|contact|person|employee|staff|user",
    re.IGNORECASE,
)
ID_COL_PATTERNS = re.compile(
    r"^id$|_id$|^ref|reference|code|number|num$|no$|ticket|order|invoice|booking",
    re.IGNORECASE,
)
REQUIRED_COL_PATTERNS = re.compile(
    r"name|email|phone|id|ref|date|status|amount|price|total|address|city|country|title|department",
    re.IGNORECASE,
)

# ─── HELPER: AUTO-DETECT COLUMN ROLES ────────────────────────────────────────

def detect_columns(df):
    """Auto-detect column roles based on name patterns and content analysis."""
    roles = {
        "email": [],
        "phone": [],
        "date": [],
        "status": [],
        "name": [],
        "id": [],
        "required": [],
        "numeric": [],
        "text": [],
    }

    for col in df.columns:
        col_str = str(col)

        # Pattern-based detection
        if EMAIL_COL_PATTERNS.search(col_str):
            roles["email"].append(col)
        if PHONE_COL_PATTERNS.search(col_str):
            # Don't match financial columns that happen to contain "num" etc.
            if not re.search(r"amount|price|bill|cost|fee|salary|revenue|total|discount", col_str, re.IGNORECASE):
                roles["phone"].append(col)
        if DATE_COL_PATTERNS.search(col_str):
            roles["date"].append(col)
        if STATUS_COL_PATTERNS.search(col_str):
            roles["status"].append(col)
        if NAME_COL_PATTERNS.search(col_str):
            roles["name"].append(col)
        if ID_COL_PATTERNS.search(col_str):
            roles["id"].append(col)
        if REQUIRED_COL_PATTERNS.search(col_str):
            roles["required"].append(col)

        # Content-based detection (fallback if no pattern match)
        if col not in roles["email"]:
            sample = df[col].dropna().head(50).astype(str)
            email_matches = sample.apply(lambda x: bool(EMAIL_REGEX.match(x.strip()))).sum()
            if email_matches > len(sample) * 0.5 and len(sample) > 0:
                roles["email"].append(col)

        if col not in roles["phone"] and col not in roles["date"]:
            # Skip numeric columns — phone numbers are stored as text
            if not pd.api.types.is_numeric_dtype(df[col]):
                sample = df[col].dropna().head(50).astype(str)
                phone_matches = sample.apply(lambda x: bool(PHONE_REGEX.match(x.strip()))).sum()
                if phone_matches > len(sample) * 0.7 and len(sample) > 0:
                    # Extra check: at least 7 digits and no date-like patterns
                    digit_counts = sample.apply(lambda x: sum(c.isdigit() for c in x))
                    has_date_sep = sample.apply(lambda x: bool(re.search(r"\d{4}[\-/]\d{2}", x))).sum()
                    if digit_counts.median() >= 7 and has_date_sep < len(sample) * 0.3:
                        roles["phone"].append(col)

        # Type-based
        if pd.api.types.is_numeric_dtype(df[col]):
            roles["numeric"].append(col)
        elif pd.api.types.is_object_dtype(df[col]):
            roles["text"].append(col)

    # Ensure required includes at least name, id, and email columns
    for key in ["name", "id", "email"]:
        for c in roles[key]:
            if c not in roles["required"]:
                roles["required"].append(c)

    return roles


# ═══════════════════════════════════════════════════════════════════════════════
# 1. REQUIRED FIELD VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════

def validate_required_fields(df, required_cols=None):
    """
    Check required columns for missing/blank values.
    Auto-detects required columns if none provided.

    Returns:
        dict with 'column', 'missing_count', 'missing_pct', 'missing_rows' per col
    """
    if required_cols is None:
        roles = detect_columns(df)
        required_cols = roles["required"]
        # If nothing detected, treat all columns as required
        if not required_cols:
            required_cols = list(df.columns)

    results = []
    for col in required_cols:
        if col not in df.columns:
            continue
        # Check for NaN, empty string, and whitespace-only
        is_missing = df[col].isna() | df[col].astype(str).str.strip().eq("")
        missing_count = int(is_missing.sum())
        missing_rows = list(df.index[is_missing])
        results.append({
            "column": col,
            "missing_count": missing_count,
            "missing_pct": round(missing_count / len(df) * 100, 2) if len(df) > 0 else 0,
            "missing_rows": missing_rows[:100],  # cap at 100 for performance
            "status": "PASS" if missing_count == 0 else "FAIL",
        })
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# 2. EMAIL VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════

def validate_emails(df, email_cols=None):
    """
    Validate email format in detected/specified columns.

    Returns:
        list of dicts with column, invalid_count, invalid_rows, sample_invalid
    """
    if email_cols is None:
        roles = detect_columns(df)
        email_cols = roles["email"]

    results = []
    for col in email_cols:
        if col not in df.columns:
            continue
        non_null = df[col].dropna().astype(str).str.strip()
        non_empty = non_null[non_null != ""]
        invalid_mask = ~non_empty.apply(lambda x: bool(EMAIL_REGEX.match(x)))
        invalid_vals = non_empty[invalid_mask]
        results.append({
            "column": col,
            "total_checked": len(non_empty),
            "valid_count": len(non_empty) - len(invalid_vals),
            "invalid_count": len(invalid_vals),
            "invalid_pct": round(len(invalid_vals) / len(non_empty) * 100, 2) if len(non_empty) > 0 else 0,
            "invalid_rows": list(invalid_vals.index[:50]),
            "sample_invalid": list(invalid_vals.head(10).values),
            "status": "PASS" if len(invalid_vals) == 0 else "FAIL",
        })
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# 3. PHONE VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════

def validate_phones(df, phone_cols=None):
    """
    Validate phone number format in detected/specified columns.
    Checks: valid pattern, minimum digit count, no letters.
    """
    if phone_cols is None:
        roles = detect_columns(df)
        phone_cols = roles["phone"]

    results = []
    for col in phone_cols:
        if col not in df.columns:
            continue
        non_null = df[col].dropna().astype(str).str.strip()
        non_empty = non_null[non_null != ""]

        def is_valid_phone(val):
            if not PHONE_REGEX.match(val):
                return False
            digit_count = sum(c.isdigit() for c in val)
            if digit_count < 7 or digit_count > 15:
                return False
            # No letters allowed
            if re.search(r"[a-zA-Z]", val):
                return False
            return True

        invalid_mask = ~non_empty.apply(is_valid_phone)
        invalid_vals = non_empty[invalid_mask]
        results.append({
            "column": col,
            "total_checked": len(non_empty),
            "valid_count": len(non_empty) - len(invalid_vals),
            "invalid_count": len(invalid_vals),
            "invalid_pct": round(len(invalid_vals) / len(non_empty) * 100, 2) if len(non_empty) > 0 else 0,
            "invalid_rows": list(invalid_vals.index[:50]),
            "sample_invalid": list(invalid_vals.head(10).values),
            "status": "PASS" if len(invalid_vals) == 0 else "FAIL",
        })
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# 4. DUPLICATE DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

def detect_duplicates(df, subset_cols=None):
    """
    Detect duplicate rows. If subset_cols given, check only those.
    Otherwise auto-detect ID/key columns, fallback to full row comparison.

    Returns:
        dict with total, duplicate_count, duplicate_groups, sample_duplicates
    """
    if subset_cols is None:
        roles = detect_columns(df)
        # Prefer ID columns, then email, then full row
        subset_cols = roles["id"] + roles["email"]
        if not subset_cols:
            subset_cols = None  # full row comparison

    if subset_cols:
        # Only use columns that actually exist
        subset_cols = [c for c in subset_cols if c in df.columns]
        if not subset_cols:
            subset_cols = None

    duplicated = df.duplicated(subset=subset_cols, keep=False)
    dup_df = df[duplicated]

    # Group duplicates
    groups = []
    if len(dup_df) > 0 and subset_cols:
        grouped = dup_df.groupby(subset_cols)
        for name, group in list(grouped)[:20]:  # max 20 groups
            groups.append({
                "key": str(name),
                "count": len(group),
                "row_indices": list(group.index[:10]),
            })
    elif len(dup_df) > 0:
        # Full row duplicates — harder to group, just show count
        first_dup = df.duplicated(keep="first")
        groups = [{"key": "full_row", "count": int(first_dup.sum()), "row_indices": list(df.index[first_dup][:20])}]

    return {
        "total_rows": len(df),
        "duplicate_rows": int(duplicated.sum()),
        "unique_duplicates": int(df.duplicated(subset=subset_cols, keep="first").sum()),
        "duplicate_pct": round(duplicated.sum() / len(df) * 100, 2) if len(df) > 0 else 0,
        "checked_columns": subset_cols or "all columns",
        "groups": groups,
        "status": "PASS" if duplicated.sum() == 0 else "WARN",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 5. ERROR DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

def detect_errors(df):
    """
    Comprehensive error detection:
      - Data type mismatches (text in numeric columns)
      - Outliers (IQR method for numeric columns)
      - Inconsistent categories (similar but not identical values)
      - Negative values where unexpected
      - Future dates
      - Impossible values
    """
    errors = []

    for col in df.columns:
        col_errors = []

        # ── Type mismatches in numeric columns ──
        if pd.api.types.is_numeric_dtype(df[col]):
            # Check for NaN that might be from failed conversion
            nan_count = int(df[col].isna().sum())
            if nan_count > 0:
                col_errors.append({
                    "type": "missing_values",
                    "count": nan_count,
                    "detail": f"{nan_count} missing/NaN values",
                })

            # Outliers (IQR method)
            valid = df[col].dropna()
            if len(valid) > 10:
                q1 = valid.quantile(0.25)
                q3 = valid.quantile(0.75)
                iqr = q3 - q1
                if iqr > 0:
                    lower = q1 - 1.5 * iqr
                    upper = q3 + 1.5 * iqr
                    outliers = valid[(valid < lower) | (valid > upper)]
                    if len(outliers) > 0:
                        col_errors.append({
                            "type": "outliers",
                            "count": len(outliers),
                            "detail": f"{len(outliers)} outliers (below {lower:.2f} or above {upper:.2f})",
                            "rows": list(outliers.index[:20]),
                        })

            # Negative values in amount/price/quantity columns
            if re.search(r"amount|price|qty|quantity|total|cost|revenue|bill|salary|fee",
                         str(col), re.IGNORECASE):
                negatives = valid[valid < 0]
                if len(negatives) > 0:
                    col_errors.append({
                        "type": "negative_values",
                        "count": len(negatives),
                        "detail": f"{len(negatives)} negative values in {col}",
                        "rows": list(negatives.index[:20]),
                    })

        # ── Date column checks ──
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            future = df[col].dropna()
            future_mask = future > pd.Timestamp.now()
            if future_mask.sum() > 0:
                col_errors.append({
                    "type": "future_dates",
                    "count": int(future_mask.sum()),
                    "detail": f"{future_mask.sum()} dates in the future",
                    "rows": list(future[future_mask].index[:20]),
                })

        # ── Text column: inconsistent categories ──
        elif pd.api.types.is_object_dtype(df[col]):
            unique_vals = df[col].dropna().unique()
            if 2 <= len(unique_vals) <= 100:
                # Check for near-duplicates (case/whitespace differences)
                normalized = pd.Series(unique_vals).str.strip().str.lower()
                norm_counts = normalized.value_counts()
                inconsistent = norm_counts[norm_counts > 1]
                if len(inconsistent) > 0:
                    for norm_val, count in inconsistent.items():
                        variants = [v for v in unique_vals if str(v).strip().lower() == norm_val]
                        col_errors.append({
                            "type": "inconsistent_category",
                            "count": count,
                            "detail": f"Inconsistent variants: {variants}",
                        })

        if col_errors:
            errors.append({"column": col, "errors": col_errors})

    total_error_count = sum(
        sum(e["count"] for e in col_info["errors"]) for col_info in errors
    )

    return {
        "total_errors": total_error_count,
        "columns_with_errors": len(errors),
        "details": errors,
        "status": "PASS" if total_error_count == 0 else "FAIL",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 6. ERROR NOTES
# ═══════════════════════════════════════════════════════════════════════════════

def generate_error_notes(validation_results):
    """
    Generate human-readable error notes from all validation results.
    Takes the combined output from run_full_validation().
    """
    notes = []

    # Required fields
    for r in validation_results.get("required_fields", []):
        if r["status"] == "FAIL":
            notes.append(f"⚠️ '{r['column']}' has {r['missing_count']} missing values ({r['missing_pct']}%)")

    # Emails
    for r in validation_results.get("email_validation", []):
        if r["status"] == "FAIL":
            samples = ", ".join(r["sample_invalid"][:3])
            notes.append(f"❌ '{r['column']}' has {r['invalid_count']} invalid emails (e.g. {samples})")

    # Phones
    for r in validation_results.get("phone_validation", []):
        if r["status"] == "FAIL":
            samples = ", ".join(r["sample_invalid"][:3])
            notes.append(f"❌ '{r['column']}' has {r['invalid_count']} invalid phone numbers (e.g. {samples})")

    # Duplicates
    dup = validation_results.get("duplicates", {})
    if dup.get("status") == "WARN":
        notes.append(f"🔁 {dup['unique_duplicates']} duplicate entries found ({dup['duplicate_pct']}% of data)")

    # Errors
    err = validation_results.get("errors", {})
    if err.get("details"):
        for col_info in err["details"]:
            for e in col_info["errors"]:
                notes.append(f"🔍 '{col_info['column']}': {e['detail']}")

    if not notes:
        notes.append("✅ No issues found — data looks clean!")

    return notes


# ═══════════════════════════════════════════════════════════════════════════════
# 7. OVERALL STATUS
# ═══════════════════════════════════════════════════════════════════════════════

def compute_overall_status(validation_results):
    """
    Compute a single overall status for the dataset.
    Returns: 'CLEAN', 'WARNINGS', or 'ERRORS'
    Plus a score 0-100 and breakdown.
    """
    issues = 0
    total_checks = 0

    # Required fields
    for r in validation_results.get("required_fields", []):
        total_checks += 1
        if r["status"] == "FAIL":
            issues += 1

    # Email
    for r in validation_results.get("email_validation", []):
        total_checks += 1
        if r["status"] == "FAIL":
            issues += 1

    # Phone
    for r in validation_results.get("phone_validation", []):
        total_checks += 1
        if r["status"] == "FAIL":
            issues += 1

    # Duplicates
    dup = validation_results.get("duplicates", {})
    total_checks += 1
    if dup.get("status") == "WARN":
        issues += 0.5  # warning = half weight

    # Errors
    err = validation_results.get("errors", {})
    total_checks += 1
    if err.get("status") == "FAIL":
        issues += 1

    score = round((1 - issues / max(total_checks, 1)) * 100, 1)

    if issues == 0:
        status = "CLEAN"
    elif issues <= 2:
        status = "WARNINGS"
    else:
        status = "ERRORS"

    return {
        "status": status,
        "score": score,
        "checks_run": total_checks,
        "issues_found": issues,
        "label": f"{score}% — {status}",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 8. WORKFLOW STATUS
# ═══════════════════════════════════════════════════════════════════════════════

def compute_workflow_status(df, status_col=None):
    """
    Analyze workflow/status distribution in the dataset.
    Auto-detects status columns if not specified.

    Returns:
        dict with distribution, completion_rate, bottleneck analysis
    """
    if status_col is None:
        roles = detect_columns(df)
        status_cols = roles["status"]
        if not status_cols:
            return {"message": "No status/workflow column detected", "status": "N/A"}
        status_col = status_cols[0]

    if status_col not in df.columns:
        return {"message": f"Column '{status_col}' not found", "status": "N/A"}

    dist = df[status_col].value_counts()
    dist_pct = df[status_col].value_counts(normalize=True) * 100

    # Detect common completion indicators
    complete_keywords = re.compile(
        r"complete|done|finished|closed|approved|delivered|resolved|paid|shipped",
        re.IGNORECASE,
    )
    pending_keywords = re.compile(
        r"pending|waiting|hold|review|processing|in[\-_\s]?progress|open|active|draft",
        re.IGNORECASE,
    )
    failed_keywords = re.compile(
        r"fail|reject|cancel|error|invalid|return|refund|declined",
        re.IGNORECASE,
    )

    completed = sum(
        int(dist.get(val, 0))
        for val in dist.index
        if complete_keywords.search(str(val))
    )
    pending = sum(
        int(dist.get(val, 0))
        for val in dist.index
        if pending_keywords.search(str(val))
    )
    failed = sum(
        int(dist.get(val, 0))
        for val in dist.index
        if failed_keywords.search(str(val))
    )

    total = len(df)
    completion_rate = round(completed / total * 100, 2) if total > 0 else 0

    return {
        "column": status_col,
        "distribution": {str(k): int(v) for k, v in dist.items()},
        "distribution_pct": {str(k): round(float(v), 2) for k, v in dist_pct.items()},
        "total": total,
        "completed": completed,
        "pending": pending,
        "failed": failed,
        "completion_rate": completion_rate,
        "bottleneck": str(dist_pct.idxmax()) if len(dist_pct) > 0 else "N/A",
        "status": "HEALTHY" if completion_rate >= 80 else "ATTENTION" if completion_rate >= 50 else "CRITICAL",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 9. AUTO COMPLETION DATE
# ═══════════════════════════════════════════════════════════════════════════════

def auto_completion_date(df, status_col=None, date_col=None):
    """
    Estimate completion date for pending items based on historical throughput.
    Uses moving average of completion rate to project future completion.
    """
    roles = detect_columns(df)

    if status_col is None:
        status_cols = roles["status"]
        if not status_cols:
            return {"message": "No status column detected", "status": "N/A"}
        status_col = status_cols[0]

    if date_col is None:
        date_cols = roles["date"]
        if not date_cols:
            return {"message": "No date column detected", "status": "N/A"}
        date_col = date_cols[0]

    if status_col not in df.columns or date_col not in df.columns:
        return {"message": "Required columns not found", "status": "N/A"}

    # Convert date column
    df_work = df.copy()
    df_work["_parsed_date"] = pd.to_datetime(df_work[date_col], errors="coerce")

    # Count completed items
    complete_keywords = re.compile(
        r"complete|done|finished|closed|approved|delivered|resolved|paid|shipped",
        re.IGNORECASE,
    )
    pending_keywords = re.compile(
        r"pending|waiting|hold|review|processing|in[\-_\s]?progress|open|active|draft",
        re.IGNORECASE,
    )

    completed_mask = df_work[status_col].astype(str).apply(lambda x: bool(complete_keywords.search(x)))
    pending_mask = df_work[status_col].astype(str).apply(lambda x: bool(pending_keywords.search(x)))

    completed_count = int(completed_mask.sum())
    pending_count = int(pending_mask.sum())

    if pending_count == 0:
        return {
            "message": "All items appear to be completed",
            "completed": completed_count,
            "pending": 0,
            "estimated_completion": "Already done",
            "status": "COMPLETE",
        }

    # Calculate throughput (items completed per day)
    completed_dates = df_work.loc[completed_mask, "_parsed_date"].dropna()
    if len(completed_dates) < 2:
        return {
            "message": "Not enough completion history to estimate",
            "completed": completed_count,
            "pending": pending_count,
            "status": "INSUFFICIENT_DATA",
        }

    date_range = (completed_dates.max() - completed_dates.min()).days
    if date_range <= 0:
        date_range = 1

    daily_throughput = completed_count / date_range
    if daily_throughput <= 0:
        daily_throughput = 0.1  # minimum estimate

    days_remaining = int(np.ceil(pending_count / daily_throughput))
    estimated_date = datetime.now() + pd.Timedelta(days=days_remaining)

    return {
        "completed": completed_count,
        "pending": pending_count,
        "daily_throughput": round(daily_throughput, 2),
        "days_remaining": days_remaining,
        "estimated_completion": estimated_date.strftime("%Y-%m-%d"),
        "status": "ESTIMATED",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 10. CONDITIONAL FORMATTING RULES
# ═══════════════════════════════════════════════════════════════════════════════

def build_conditional_formatting(df, validation_results):
    """
    Build conditional formatting rules based on validation results.
    Returns rules that can be applied in Streamlit UI or Excel export.

    Each rule: {column, condition, style, rows}
    """
    rules = []

    # Missing required fields → red background
    for r in validation_results.get("required_fields", []):
        if r["status"] == "FAIL" and r["missing_rows"]:
            rules.append({
                "column": r["column"],
                "condition": "is_blank",
                "color": "#FFCCCC",       # light red
                "font_color": "#CC0000",  # dark red
                "label": "Missing required value",
                "rows": r["missing_rows"],
            })

    # Invalid emails → orange
    for r in validation_results.get("email_validation", []):
        if r["status"] == "FAIL" and r["invalid_rows"]:
            rules.append({
                "column": r["column"],
                "condition": "invalid_email",
                "color": "#FFE0CC",       # light orange
                "font_color": "#CC6600",
                "label": "Invalid email format",
                "rows": r["invalid_rows"],
            })

    # Invalid phones → yellow
    for r in validation_results.get("phone_validation", []):
        if r["status"] == "FAIL" and r["invalid_rows"]:
            rules.append({
                "column": r["column"],
                "condition": "invalid_phone",
                "color": "#FFFFCC",       # light yellow
                "font_color": "#999900",
                "label": "Invalid phone number",
                "rows": r["invalid_rows"],
            })

    # Duplicates → purple tint
    dup = validation_results.get("duplicates", {})
    if dup.get("status") == "WARN":
        dup_rows = []
        for g in dup.get("groups", []):
            dup_rows.extend(g.get("row_indices", []))
        if dup_rows:
            rules.append({
                "column": "__all__",
                "condition": "duplicate_row",
                "color": "#E8CCF5",       # light purple
                "font_color": "#7700CC",
                "label": "Duplicate entry",
                "rows": dup_rows,
            })

    # Outliers → blue tint
    err = validation_results.get("errors", {})
    for col_info in err.get("details", []):
        for e in col_info.get("errors", []):
            if e["type"] == "outliers":
                rules.append({
                    "column": col_info["column"],
                    "condition": "outlier",
                    "color": "#CCE0FF",   # light blue
                    "font_color": "#0044CC",
                    "label": "Statistical outlier",
                    "rows": e.get("rows", []),
                })
            elif e["type"] == "negative_values":
                rules.append({
                    "column": col_info["column"],
                    "condition": "negative",
                    "color": "#FFCCCC",
                    "font_color": "#CC0000",
                    "label": "Negative value",
                    "rows": e.get("rows", []),
                })

    return rules


# ═══════════════════════════════════════════════════════════════════════════════
# 11. ERROR RATE CALCULATION
# ═══════════════════════════════════════════════════════════════════════════════

def calculate_error_rate(df, validation_results):
    """
    Calculate overall and per-column error rates.
    Error rate = (cells with issues / total cells) * 100
    """
    total_cells = df.shape[0] * df.shape[1]
    error_cells = 0
    per_column = {}

    # Missing values
    for r in validation_results.get("required_fields", []):
        count = r["missing_count"]
        error_cells += count
        per_column[r["column"]] = per_column.get(r["column"], 0) + count

    # Invalid emails
    for r in validation_results.get("email_validation", []):
        count = r["invalid_count"]
        error_cells += count
        per_column[r["column"]] = per_column.get(r["column"], 0) + count

    # Invalid phones
    for r in validation_results.get("phone_validation", []):
        count = r["invalid_count"]
        error_cells += count
        per_column[r["column"]] = per_column.get(r["column"], 0) + count

    # Data errors
    err = validation_results.get("errors", {})
    for col_info in err.get("details", []):
        for e in col_info.get("errors", []):
            error_cells += e["count"]
            per_column[col_info["column"]] = per_column.get(col_info["column"], 0) + e["count"]

    overall_rate = round(error_cells / max(total_cells, 1) * 100, 4)

    # Per-column rates
    column_rates = {}
    for col, count in per_column.items():
        col_total = len(df)
        column_rates[col] = {
            "error_count": count,
            "error_rate": round(count / max(col_total, 1) * 100, 2),
        }

    return {
        "total_cells": total_cells,
        "error_cells": error_cells,
        "overall_error_rate": overall_rate,
        "per_column": column_rates,
        "cleanest_columns": [c for c in df.columns if c not in per_column],
        "worst_column": max(column_rates, key=lambda k: column_rates[k]["error_rate"]) if column_rates else None,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 12. DATA ACCURACY CALCULATION
# ═══════════════════════════════════════════════════════════════════════════════

def calculate_data_accuracy(df, validation_results):
    """
    Data accuracy = 100% - error_rate.
    Also provides completeness, consistency, and validity sub-scores.
    """
    error_rates = calculate_error_rate(df, validation_results)

    total_cells = df.shape[0] * df.shape[1]

    # Completeness: % of non-missing cells
    missing_cells = int(df.isna().sum().sum())
    # Also count blank strings
    blank_cells = 0
    for col in df.select_dtypes(include="object").columns:
        blank_cells += int(df[col].astype(str).str.strip().eq("").sum())
    completeness = round((1 - (missing_cells + blank_cells) / max(total_cells, 1)) * 100, 2)

    # Validity: % of cells passing format validation (email, phone)
    validity_errors = 0
    validity_checked = 0
    for r in validation_results.get("email_validation", []):
        validity_checked += r["total_checked"]
        validity_errors += r["invalid_count"]
    for r in validation_results.get("phone_validation", []):
        validity_checked += r["total_checked"]
        validity_errors += r["invalid_count"]
    validity = round((1 - validity_errors / max(validity_checked, 1)) * 100, 2) if validity_checked > 0 else 100.0

    # Consistency: based on duplicate rate and category inconsistencies
    dup = validation_results.get("duplicates", {})
    dup_rate = dup.get("duplicate_pct", 0)
    consistency = round(100 - dup_rate, 2)

    # Overall accuracy
    accuracy = round((completeness + validity + consistency) / 3, 2)

    return {
        "overall_accuracy": accuracy,
        "completeness": completeness,
        "validity": validity,
        "consistency": consistency,
        "error_rate": error_rates["overall_error_rate"],
        "label": f"{accuracy}% accurate",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 13. PRODUCTIVITY TRACKER
# ═══════════════════════════════════════════════════════════════════════════════

def track_productivity(df, date_col=None, group_col=None, value_col=None):
    """
    Track productivity metrics:
      - Items processed per day/week/month
      - Trends over time
      - Per-group breakdown (if group_col provided)

    Auto-detects columns if not specified.
    """
    roles = detect_columns(df)

    if date_col is None:
        date_cols = roles["date"]
        if not date_cols:
            return {"message": "No date column detected for productivity tracking", "status": "N/A"}
        date_col = date_cols[0]

    if date_col not in df.columns:
        return {"message": f"Column '{date_col}' not found", "status": "N/A"}

    df_work = df.copy()
    df_work["_date"] = pd.to_datetime(df_work[date_col], errors="coerce")
    df_work = df_work.dropna(subset=["_date"])

    if len(df_work) == 0:
        return {"message": "No valid dates found", "status": "N/A"}

    # Daily counts
    daily = df_work.groupby(df_work["_date"].dt.date).size()
    weekly = df_work.groupby(df_work["_date"].dt.isocalendar().week).size()
    monthly = df_work.groupby(df_work["_date"].dt.to_period("M")).size()

    result = {
        "date_column": date_col,
        "total_records": len(df_work),
        "date_range": f"{df_work['_date'].min().strftime('%Y-%m-%d')} to {df_work['_date'].max().strftime('%Y-%m-%d')}",
        "days_span": (df_work["_date"].max() - df_work["_date"].min()).days + 1,
        "daily_avg": round(daily.mean(), 2),
        "daily_max": int(daily.max()),
        "daily_min": int(daily.min()),
        "weekly_avg": round(weekly.mean(), 2),
        "monthly_avg": round(monthly.mean(), 2),
        "peak_day": str(daily.idxmax()),
        "peak_count": int(daily.max()),
        "daily_trend": {str(k): int(v) for k, v in daily.tail(30).items()},
        "monthly_trend": {str(k): int(v) for k, v in monthly.items()},
        "status": "OK",
    }

    # Per-group breakdown
    if group_col and group_col in df.columns:
        group_prod = df_work.groupby(group_col).agg(
            count=("_date", "size"),
            first_date=("_date", "min"),
            last_date=("_date", "max"),
        )
        group_prod["days_active"] = (group_prod["last_date"] - group_prod["first_date"]).dt.days + 1
        group_prod["daily_avg"] = (group_prod["count"] / group_prod["days_active"]).round(2)
        group_prod = group_prod.sort_values("count", ascending=False)

        result["group_column"] = group_col
        result["group_breakdown"] = {
            str(idx): {
                "count": int(row["count"]),
                "daily_avg": float(row["daily_avg"]),
                "days_active": int(row["days_active"]),
            }
            for idx, row in group_prod.head(20).iterrows()
        }
        result["top_performer"] = str(group_prod.index[0]) if len(group_prod) > 0 else "N/A"

    elif group_col is None:
        # Auto-detect a reasonable grouping column
        candidates = roles["name"] + roles["status"]
        for c in candidates:
            if c in df.columns and 2 <= df[c].nunique() <= 50:
                group_col = c
                break
        if group_col:
            return track_productivity(df, date_col=date_col, group_col=group_col, value_col=value_col)

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# MASTER: RUN FULL VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════

def run_full_validation(df, config=None):
    """
    Run all 13 validation functions on a dataframe.

    Args:
        df: pandas DataFrame
        config: optional dict to override auto-detection:
            {
                'required_cols': [...],
                'email_cols': [...],
                'phone_cols': [...],
                'status_col': 'Status',
                'date_col': 'Date',
                'group_col': 'Department',
                'duplicate_cols': [...],
            }

    Returns:
        dict with all validation results keyed by function name
    """
    config = config or {}
    results = {}

    # 1. Required fields
    results["required_fields"] = validate_required_fields(
        df, required_cols=config.get("required_cols")
    )

    # 2. Email validation
    results["email_validation"] = validate_emails(
        df, email_cols=config.get("email_cols")
    )

    # 3. Phone validation
    results["phone_validation"] = validate_phones(
        df, phone_cols=config.get("phone_cols")
    )

    # 4. Duplicate detection
    results["duplicates"] = detect_duplicates(
        df, subset_cols=config.get("duplicate_cols")
    )

    # 5. Error detection
    results["errors"] = detect_errors(df)

    # 6. Error notes
    results["error_notes"] = generate_error_notes(results)

    # 7. Overall status
    results["overall_status"] = compute_overall_status(results)

    # 8. Workflow status
    results["workflow_status"] = compute_workflow_status(
        df, status_col=config.get("status_col")
    )

    # 9. Auto completion date
    results["auto_completion"] = auto_completion_date(
        df,
        status_col=config.get("status_col"),
        date_col=config.get("date_col"),
    )

    # 10. Conditional formatting
    results["conditional_formatting"] = build_conditional_formatting(df, results)

    # 11. Error rate
    results["error_rate"] = calculate_error_rate(df, results)

    # 12. Data accuracy
    results["data_accuracy"] = calculate_data_accuracy(df, results)

    # 13. Productivity tracker
    results["productivity"] = track_productivity(
        df,
        date_col=config.get("date_col"),
        group_col=config.get("group_col"),
    )

    # Detected columns (for reference)
    results["detected_columns"] = detect_columns(df)

    return results


# ═══════════════════════════════════════════════════════════════════════════════
# SUMMARY: Quick text report
# ═══════════════════════════════════════════════════════════════════════════════

def validation_summary_text(results):
    """Generate a concise text summary of validation results."""
    lines = []
    overall = results.get("overall_status", {})
    accuracy = results.get("data_accuracy", {})

    lines.append(f"📊 Data Quality Score: {overall.get('score', 'N/A')}% — {overall.get('status', 'N/A')}")
    lines.append(f"🎯 Data Accuracy: {accuracy.get('overall_accuracy', 'N/A')}%")
    lines.append(f"   Completeness: {accuracy.get('completeness', 'N/A')}%")
    lines.append(f"   Validity: {accuracy.get('validity', 'N/A')}%")
    lines.append(f"   Consistency: {accuracy.get('consistency', 'N/A')}%")

    err_rate = results.get("error_rate", {})
    lines.append(f"❌ Error Rate: {err_rate.get('overall_error_rate', 'N/A')}%")
    if err_rate.get("worst_column"):
        wc = err_rate["worst_column"]
        lines.append(f"   Worst column: {wc} ({err_rate['per_column'][wc]['error_rate']}% errors)")

    workflow = results.get("workflow_status", {})
    if workflow.get("column"):
        lines.append(f"📋 Workflow: {workflow.get('completion_rate', 'N/A')}% completed ({workflow['column']})")
        lines.append(f"   Bottleneck: {workflow.get('bottleneck', 'N/A')}")

    completion = results.get("auto_completion", {})
    if completion.get("estimated_completion") and completion["status"] == "ESTIMATED":
        lines.append(f"📅 Est. completion: {completion['estimated_completion']} ({completion['days_remaining']} days)")

    prod = results.get("productivity", {})
    if prod.get("status") == "OK":
        lines.append(f"⚡ Productivity: {prod['daily_avg']}/day avg, peak {prod['peak_count']} on {prod['peak_day']}")

    lines.append("")
    lines.append("── Issues ──")
    for note in results.get("error_notes", []):
        lines.append(f"  {note}")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT: Validation Report with Dashboard
# ═══════════════════════════════════════════════════════════════════════════════

def export_validation_report(df, results, output_path="validation_report.xlsx"):
    """
    Export validation results as a multi-sheet Excel workbook.

    Sheets:
      1. Dashboard  — summary cards, scores, key metrics
      2. Original Data — with conditional formatting applied
      3. Validation Summary — scores and check results table
      4. Error Details — every issue per row/column
      5. Duplicates — duplicate groups (if any)
      6. Workflow & Productivity — status distribution + trends (if applicable)
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt

    wb = Workbook()

    # ── Colour palette ──
    DARK_BG     = "0F172A"
    CARD_BG     = "1E293B"
    WHITE       = "FFFFFF"
    LIGHT_GRAY  = "F1F5F9"
    MID_GRAY    = "94A3B8"
    GREEN       = "16A34A"
    GREEN_BG    = "DCFCE7"
    BLUE        = "2563EB"
    BLUE_BG     = "DBEAFE"
    RED         = "DC2626"
    RED_BG      = "FEE2E2"
    ORANGE      = "D97706"
    ORANGE_BG   = "FEF3C7"
    PURPLE      = "7C3AED"
    PURPLE_BG   = "EDE9FE"
    HEADER_BG   = "1E3A5F"

    f_title     = Font(name="Arial", size=18, bold=True, color=DARK_BG)
    f_subtitle  = Font(name="Arial", size=11, color=MID_GRAY)
    f_header    = Font(name="Arial", size=11, bold=True, color=WHITE)
    f_body      = Font(name="Arial", size=10, color=DARK_BG)
    f_body_bold = Font(name="Arial", size=10, bold=True, color=DARK_BG)
    f_card_val  = Font(name="Arial", size=22, bold=True)
    f_card_lbl  = Font(name="Arial", size=9, color=MID_GRAY)
    f_card_sub  = Font(name="Arial", size=11, bold=True)
    fill_header = PatternFill("solid", fgColor=HEADER_BG)
    fill_light  = PatternFill("solid", fgColor=LIGHT_GRAY)
    align_c     = Alignment(horizontal="center", vertical="center")
    align_l     = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def write_header_row(ws, row, headers, col_start=1):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=col_start + i, value=h)
            c.font = f_header
            c.fill = fill_header
            c.alignment = align_c
            c.border = thin_border

    def write_data_row(ws, row, values, col_start=1, stripe=False):
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=col_start + i, value=v)
            c.font = f_body
            c.alignment = align_l
            c.border = thin_border
            if stripe:
                c.fill = fill_light

    # ── Extract results ──
    overall = results.get("overall_status", {})
    accuracy = results.get("data_accuracy", {})
    err_rate = results.get("error_rate", {})
    workflow = results.get("workflow_status", {})
    prod = results.get("productivity", {})
    completion = results.get("auto_completion", {})
    notes = results.get("error_notes", [])

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_properties.tabColor = BLUE
    set_col_widths(ws_dash, [3, 22, 22, 22, 22, 22, 3])

    # Title
    ws_dash.merge_cells("B2:F2")
    c = ws_dash["B2"]
    c.value = "DATA VALIDATION REPORT"
    c.font = f_title
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws_dash.merge_cells("B3:F3")
    c = ws_dash["B3"]
    c.value = f"Generated: {dt.now().strftime('%B %d, %Y at %H:%M')}  |  {df.shape[0]} rows × {df.shape[1]} columns"
    c.font = f_subtitle

    # ── Score cards (row 5-8) ──
    card_data = [
        ("Quality Score", f"{overall.get('score', 'N/A')}%", overall.get("status", "N/A"),
         GREEN if overall.get("status") == "CLEAN" else ORANGE if overall.get("status") == "WARNINGS" else RED,
         GREEN_BG if overall.get("status") == "CLEAN" else ORANGE_BG if overall.get("status") == "WARNINGS" else RED_BG),
        ("Data Accuracy", f"{accuracy.get('overall_accuracy', 'N/A')}%", "ACCURACY", BLUE, BLUE_BG),
        ("Error Rate", f"{err_rate.get('overall_error_rate', 'N/A')}%", "ERRORS", RED, RED_BG),
        ("Workflow", f"{workflow.get('completion_rate', 'N/A')}%" if workflow.get("column") else "N/A",
         "COMPLETION" if workflow.get("column") else "N/A", PURPLE, PURPLE_BG),
    ]

    for i, (label, value, sub_label, color, bg_color) in enumerate(card_data):
        col = 2 + i
        # Card value
        c = ws_dash.cell(row=5, column=col, value=value)
        c.font = Font(name="Arial", size=22, bold=True, color=color)
        c.fill = PatternFill("solid", fgColor=bg_color)
        c.alignment = align_c
        c.border = thin_border
        # Card sublabel
        c = ws_dash.cell(row=6, column=col, value=sub_label)
        c.font = Font(name="Arial", size=9, bold=True, color=color)
        c.fill = PatternFill("solid", fgColor=bg_color)
        c.alignment = align_c
        c.border = thin_border
        # Card title
        c = ws_dash.cell(row=7, column=col, value=label)
        c.font = f_card_lbl
        c.fill = PatternFill("solid", fgColor=bg_color)
        c.alignment = align_c
        c.border = thin_border

    # Blank row 8 as spacer — just set height
    ws_dash.row_dimensions[8].height = 8

    # ── Accuracy Breakdown (row 9-11) ──
    ws_dash.merge_cells("B9:E9")
    c = ws_dash["B9"]
    c.value = "ACCURACY BREAKDOWN"
    c.font = Font(name="Arial", size=12, bold=True, color=DARK_BG)

    sub_scores = [
        ("Completeness", f"{accuracy.get('completeness', 'N/A')}%"),
        ("Validity", f"{accuracy.get('validity', 'N/A')}%"),
        ("Consistency", f"{accuracy.get('consistency', 'N/A')}%"),
    ]
    for i, (lbl, val) in enumerate(sub_scores):
        col = 2 + i
        c = ws_dash.cell(row=10, column=col, value=val)
        c.font = Font(name="Arial", size=16, bold=True, color=BLUE)
        c.fill = PatternFill("solid", fgColor=BLUE_BG)
        c.alignment = align_c
        c.border = thin_border
        c = ws_dash.cell(row=11, column=col, value=lbl)
        c.font = f_card_lbl
        c.fill = PatternFill("solid", fgColor=BLUE_BG)
        c.alignment = align_c
        c.border = thin_border

    # ── Data Profile (row 13+) ──
    ws_dash.merge_cells("B13:E13")
    c = ws_dash["B13"]
    c.value = "DATA PROFILE"
    c.font = Font(name="Arial", size=12, bold=True, color=DARK_BG)

    profile_data = [
        ("Total Rows", df.shape[0]),
        ("Total Columns", df.shape[1]),
        ("Total Cells", df.shape[0] * df.shape[1]),
        ("Missing Values", int(df.isna().sum().sum())),
        ("Duplicate Rows", results.get("duplicates", {}).get("duplicate_rows", 0)),
        ("Error Cells", err_rate.get("error_cells", 0)),
    ]
    write_header_row(ws_dash, 14, ["Metric", "Value"], col_start=2)
    for i, (metric, value) in enumerate(profile_data):
        r = 15 + i
        write_data_row(ws_dash, r, [metric, value], col_start=2, stripe=i % 2 == 0)

    # ── Issues Found (row 22+) ──
    issue_start = 15 + len(profile_data) + 1
    ws_dash.merge_cells(f"B{issue_start}:E{issue_start}")
    c = ws_dash.cell(row=issue_start, column=2, value="ISSUES FOUND")
    c.font = Font(name="Arial", size=12, bold=True, color=DARK_BG)

    write_header_row(ws_dash, issue_start + 1, ["#", "Issue"], col_start=2)
    for i, note in enumerate(notes):
        # Strip emoji for cleaner Excel
        clean_note = note.replace("⚠️", "[WARN]").replace("❌", "[FAIL]").replace("🔁", "[DUP]").replace("🔍", "[CHECK]").replace("✅", "[OK]")
        write_data_row(ws_dash, issue_start + 2 + i, [i + 1, clean_note], col_start=2, stripe=i % 2 == 0)

    # ── Workflow summary on dashboard (if applicable) ──
    wf_start = issue_start + 2 + len(notes) + 1
    if workflow.get("column"):
        ws_dash.merge_cells(f"B{wf_start}:E{wf_start}")
        c = ws_dash.cell(row=wf_start, column=2, value="WORKFLOW STATUS")
        c.font = Font(name="Arial", size=12, bold=True, color=DARK_BG)

        wf_info = [
            ("Status Column", workflow["column"]),
            ("Completion Rate", f"{workflow.get('completion_rate', 'N/A')}%"),
            ("Completed", workflow.get("completed", 0)),
            ("Pending", workflow.get("pending", 0)),
            ("Failed", workflow.get("failed", 0)),
            ("Bottleneck", workflow.get("bottleneck", "N/A")),
        ]
        write_header_row(ws_dash, wf_start + 1, ["Metric", "Value"], col_start=2)
        for i, (m, v) in enumerate(wf_info):
            write_data_row(ws_dash, wf_start + 2 + i, [m, v], col_start=2, stripe=i % 2 == 0)

        # Completion estimate
        if completion.get("status") == "ESTIMATED":
            est_row = wf_start + 2 + len(wf_info)
            write_data_row(ws_dash, est_row, ["Est. Completion Date", completion["estimated_completion"]], col_start=2)
            write_data_row(ws_dash, est_row + 1, ["Days Remaining", completion["days_remaining"]], col_start=2, stripe=True)
            write_data_row(ws_dash, est_row + 2, ["Daily Throughput", completion["daily_throughput"]], col_start=2)

    # ── Productivity on dashboard (if applicable) ──
    if prod.get("status") == "OK":
        prod_start = ws_dash.max_row + 2
        ws_dash.merge_cells(f"B{prod_start}:E{prod_start}")
        c = ws_dash.cell(row=prod_start, column=2, value="PRODUCTIVITY")
        c.font = Font(name="Arial", size=12, bold=True, color=DARK_BG)

        prod_info = [
            ("Date Range", prod.get("date_range", "N/A")),
            ("Total Records", prod.get("total_records", 0)),
            ("Daily Average", prod.get("daily_avg", 0)),
            ("Peak Day", f"{prod.get('peak_day', 'N/A')} ({prod.get('peak_count', 0)} items)"),
            ("Monthly Average", prod.get("monthly_avg", 0)),
        ]
        write_header_row(ws_dash, prod_start + 1, ["Metric", "Value"], col_start=2)
        for i, (m, v) in enumerate(prod_info):
            write_data_row(ws_dash, prod_start + 2 + i, [m, v], col_start=2, stripe=i % 2 == 0)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: ORIGINAL DATA with conditional formatting
    # ══════════════════════════════════════════════════════════════════════════
    ws_data = wb.create_sheet("Original Data")
    ws_data.sheet_properties.tabColor = "334155"

    # Write headers
    for j, col_name in enumerate(df.columns, 1):
        c = ws_data.cell(row=1, column=j, value=str(col_name))
        c.font = f_header
        c.fill = fill_header
        c.alignment = align_c
        c.border = thin_border

    # Write data
    for i, (_, row_data) in enumerate(df.iterrows(), 2):
        for j, val in enumerate(row_data, 1):
            c = ws_data.cell(row=i, column=j)
            if pd.isna(val):
                c.value = None
            else:
                c.value = val
            c.font = f_body
            c.border = thin_border
            if (i - 2) % 2 == 0:
                c.fill = fill_light

    # Apply conditional formatting colors
    cf_rules = results.get("conditional_formatting", [])
    col_index_map = {str(col): idx + 1 for idx, col in enumerate(df.columns)}
    for rule in cf_rules:
        fill = PatternFill("solid", fgColor=rule["color"].replace("#", ""))
        font_c = Font(name="Arial", size=10, color=rule["font_color"].replace("#", ""))
        if rule["column"] == "__all__":
            cols_to_mark = list(range(1, len(df.columns) + 1))
        else:
            col_idx = col_index_map.get(rule["column"])
            if not col_idx:
                continue
            cols_to_mark = [col_idx]

        for row_idx in rule.get("rows", []):
            excel_row = row_idx + 2  # +1 for header, +1 for 0-indexing
            if excel_row > df.shape[0] + 1:
                continue
            for col_idx in cols_to_mark:
                cell = ws_data.cell(row=excel_row, column=col_idx)
                cell.fill = fill
                cell.font = font_c

    # Auto-fit column widths
    for j in range(1, len(df.columns) + 1):
        max_len = len(str(df.columns[j - 1]))
        for i in range(2, min(len(df) + 2, 102)):
            val = ws_data.cell(row=i, column=j).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 40))
        ws_data.column_dimensions[get_column_letter(j)].width = max_len + 4

    # Freeze top row
    ws_data.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: VALIDATION SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    ws_sum = wb.create_sheet("Validation Summary")
    ws_sum.sheet_properties.tabColor = GREEN
    set_col_widths(ws_sum, [25, 20, 18, 18, 30])

    ws_sum.merge_cells("A1:E1")
    c = ws_sum["A1"]
    c.value = "VALIDATION CHECK RESULTS"
    c.font = f_title

    # Required fields
    r = 3
    ws_sum.merge_cells(f"A{r}:E{r}")
    ws_sum.cell(row=r, column=1, value="REQUIRED FIELD CHECKS").font = Font(name="Arial", size=12, bold=True, color=DARK_BG)
    r += 1
    write_header_row(ws_sum, r, ["Column", "Missing Count", "Missing %", "Status", "Notes"])
    r += 1
    for i, req in enumerate(results.get("required_fields", [])):
        note = ""
        if req["missing_count"] > 0:
            note = f"Rows: {', '.join(str(x) for x in req['missing_rows'][:10])}"
            if len(req["missing_rows"]) > 10:
                note += f" ... (+{len(req['missing_rows']) - 10} more)"
        write_data_row(ws_sum, r, [
            req["column"], req["missing_count"], f"{req['missing_pct']}%", req["status"], note
        ], stripe=i % 2 == 0)
        # Color status cell
        status_cell = ws_sum.cell(row=r, column=4)
        if req["status"] == "PASS":
            status_cell.font = Font(name="Arial", size=10, bold=True, color=GREEN)
        else:
            status_cell.font = Font(name="Arial", size=10, bold=True, color=RED)
        r += 1

    # Email validation
    email_results = results.get("email_validation", [])
    if email_results:
        r += 1
        ws_sum.merge_cells(f"A{r}:E{r}")
        ws_sum.cell(row=r, column=1, value="EMAIL VALIDATION").font = Font(name="Arial", size=12, bold=True, color=DARK_BG)
        r += 1
        write_header_row(ws_sum, r, ["Column", "Valid", "Invalid", "Status", "Invalid Samples"])
        r += 1
        for i, ev in enumerate(email_results):
            samples = ", ".join(ev.get("sample_invalid", [])[:5])
            write_data_row(ws_sum, r, [
                ev["column"], ev["valid_count"], ev["invalid_count"], ev["status"], samples
            ], stripe=i % 2 == 0)
            status_cell = ws_sum.cell(row=r, column=4)
            status_cell.font = Font(name="Arial", size=10, bold=True, color=GREEN if ev["status"] == "PASS" else RED)
            r += 1

    # Phone validation
    phone_results = results.get("phone_validation", [])
    if phone_results:
        r += 1
        ws_sum.merge_cells(f"A{r}:E{r}")
        ws_sum.cell(row=r, column=1, value="PHONE VALIDATION").font = Font(name="Arial", size=12, bold=True, color=DARK_BG)
        r += 1
        write_header_row(ws_sum, r, ["Column", "Valid", "Invalid", "Status", "Invalid Samples"])
        r += 1
        for i, pv in enumerate(phone_results):
            samples = ", ".join(pv.get("sample_invalid", [])[:5])
            write_data_row(ws_sum, r, [
                pv["column"], pv["valid_count"], pv["invalid_count"], pv["status"], samples
            ], stripe=i % 2 == 0)
            status_cell = ws_sum.cell(row=r, column=4)
            status_cell.font = Font(name="Arial", size=10, bold=True, color=GREEN if pv["status"] == "PASS" else RED)
            r += 1

    # Per-column error rates
    col_rates = err_rate.get("per_column", {})
    if col_rates:
        r += 1
        ws_sum.merge_cells(f"A{r}:E{r}")
        ws_sum.cell(row=r, column=1, value="PER-COLUMN ERROR RATES").font = Font(name="Arial", size=12, bold=True, color=DARK_BG)
        r += 1
        write_header_row(ws_sum, r, ["Column", "Error Count", "Error Rate", "", ""])
        r += 1
        sorted_rates = sorted(col_rates.items(), key=lambda x: x[1]["error_rate"], reverse=True)
        for i, (col_name, info) in enumerate(sorted_rates):
            write_data_row(ws_sum, r, [col_name, info["error_count"], f"{info['error_rate']}%", "", ""], stripe=i % 2 == 0)
            r += 1

    ws_sum.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4: ERROR DETAILS
    # ══════════════════════════════════════════════════════════════════════════
    ws_err = wb.create_sheet("Error Details")
    ws_err.sheet_properties.tabColor = RED
    set_col_widths(ws_err, [10, 22, 22, 40, 25])

    ws_err.merge_cells("A1:E1")
    ws_err.cell(row=1, column=1, value="DETAILED ERROR LOG").font = f_title

    write_header_row(ws_err, 3, ["Row #", "Column", "Error Type", "Detail", "Current Value"])
    r = 4
    error_count = 0

    # Collect all errors with row-level detail
    for req in results.get("required_fields", []):
        if req["status"] == "FAIL":
            for row_idx in req["missing_rows"][:200]:
                write_data_row(ws_err, r, [
                    row_idx + 1, req["column"], "Missing Required", "Required field is blank/empty",
                    str(df.iloc[row_idx][req["column"]]) if row_idx < len(df) else ""
                ], stripe=error_count % 2 == 0)
                error_count += 1
                r += 1

    for ev in results.get("email_validation", []):
        if ev["status"] == "FAIL":
            for row_idx in ev["invalid_rows"][:200]:
                val = str(df.iloc[row_idx][ev["column"]]) if row_idx < len(df) else ""
                write_data_row(ws_err, r, [
                    row_idx + 1, ev["column"], "Invalid Email", "Email format does not match pattern", val
                ], stripe=error_count % 2 == 0)
                error_count += 1
                r += 1

    for pv in results.get("phone_validation", []):
        if pv["status"] == "FAIL":
            for row_idx in pv["invalid_rows"][:200]:
                val = str(df.iloc[row_idx][pv["column"]]) if row_idx < len(df) else ""
                write_data_row(ws_err, r, [
                    row_idx + 1, pv["column"], "Invalid Phone", "Phone number format is invalid", val
                ], stripe=error_count % 2 == 0)
                error_count += 1
                r += 1

    err_details = results.get("errors", {}).get("details", [])
    for col_info in err_details:
        for e in col_info["errors"]:
            for row_idx in e.get("rows", [])[:200]:
                val = str(df.iloc[row_idx][col_info["column"]]) if row_idx < len(df) else ""
                write_data_row(ws_err, r, [
                    row_idx + 1, col_info["column"], e["type"].replace("_", " ").title(),
                    e["detail"], val
                ], stripe=error_count % 2 == 0)
                error_count += 1
                r += 1

    if error_count == 0:
        ws_err.cell(row=4, column=1, value="No errors found").font = Font(name="Arial", size=11, color=GREEN, italic=True)

    ws_err.freeze_panes = "A4"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5: DUPLICATES
    # ══════════════════════════════════════════════════════════════════════════
    dup = results.get("duplicates", {})
    ws_dup = wb.create_sheet("Duplicates")
    ws_dup.sheet_properties.tabColor = PURPLE
    set_col_widths(ws_dup, [25, 18, 40])

    ws_dup.merge_cells("A1:C1")
    ws_dup.cell(row=1, column=1, value="DUPLICATE DETECTION").font = f_title

    dup_info = [
        ("Total Rows", dup.get("total_rows", 0)),
        ("Duplicate Rows", dup.get("duplicate_rows", 0)),
        ("Duplicate %", f"{dup.get('duplicate_pct', 0)}%"),
        ("Checked Columns", str(dup.get("checked_columns", "all columns"))),
        ("Status", dup.get("status", "N/A")),
    ]
    write_header_row(ws_dup, 3, ["Metric", "Value", ""])
    for i, (m, v) in enumerate(dup_info):
        write_data_row(ws_dup, 4 + i, [m, v, ""], stripe=i % 2 == 0)

    groups = dup.get("groups", [])
    if groups:
        gr = 4 + len(dup_info) + 1
        ws_dup.merge_cells(f"A{gr}:C{gr}")
        ws_dup.cell(row=gr, column=1, value="DUPLICATE GROUPS").font = Font(name="Arial", size=12, bold=True, color=DARK_BG)
        gr += 1
        write_header_row(ws_dup, gr, ["Key Value", "Count", "Row Numbers"])
        gr += 1
        for i, g in enumerate(groups[:50]):
            rows_str = ", ".join(str(x + 1) for x in g.get("row_indices", [])[:20])
            write_data_row(ws_dup, gr, [str(g["key"]), g["count"], rows_str], stripe=i % 2 == 0)
            gr += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 6: WORKFLOW & PRODUCTIVITY
    # ══════════════════════════════════════════════════════════════════════════
    ws_wf = wb.create_sheet("Workflow & Productivity")
    ws_wf.sheet_properties.tabColor = ORANGE
    set_col_widths(ws_wf, [25, 18, 18, 18])

    ws_wf.merge_cells("A1:D1")
    ws_wf.cell(row=1, column=1, value="WORKFLOW & PRODUCTIVITY").font = f_title

    r = 3
    if workflow.get("column"):
        ws_wf.merge_cells(f"A{r}:D{r}")
        ws_wf.cell(row=r, column=1, value="WORKFLOW STATUS DISTRIBUTION").font = Font(name="Arial", size=12, bold=True, color=DARK_BG)
        r += 1
        write_header_row(ws_wf, r, ["Status Value", "Count", "Percentage", ""])
        r += 1
        dist = workflow.get("distribution", {})
        dist_pct = workflow.get("distribution_pct", {})
        for i, (status_val, count) in enumerate(dist.items()):
            pct = dist_pct.get(status_val, 0)
            write_data_row(ws_wf, r, [status_val, count, f"{pct}%", ""], stripe=i % 2 == 0)
            r += 1

        r += 1
        summary_items = [
            ("Completed", workflow.get("completed", 0)),
            ("Pending", workflow.get("pending", 0)),
            ("Failed", workflow.get("failed", 0)),
            ("Completion Rate", f"{workflow.get('completion_rate', 'N/A')}%"),
            ("Bottleneck", workflow.get("bottleneck", "N/A")),
        ]
        write_header_row(ws_wf, r, ["Metric", "Value", "", ""])
        r += 1
        for i, (m, v) in enumerate(summary_items):
            write_data_row(ws_wf, r, [m, v, "", ""], stripe=i % 2 == 0)
            r += 1

        if completion.get("status") == "ESTIMATED":
            r += 1
            ws_wf.merge_cells(f"A{r}:D{r}")
            ws_wf.cell(row=r, column=1, value="COMPLETION ESTIMATE").font = Font(name="Arial", size=12, bold=True, color=DARK_BG)
            r += 1
            est_items = [
                ("Estimated Completion Date", completion["estimated_completion"]),
                ("Days Remaining", completion["days_remaining"]),
                ("Daily Throughput", f"{completion['daily_throughput']} items/day"),
            ]
            write_header_row(ws_wf, r, ["Metric", "Value", "", ""])
            r += 1
            for i, (m, v) in enumerate(est_items):
                write_data_row(ws_wf, r, [m, v, "", ""], stripe=i % 2 == 0)
                r += 1
    else:
        ws_wf.cell(row=r, column=1, value="No workflow/status column detected").font = Font(name="Arial", size=11, color=MID_GRAY, italic=True)
        r += 2

    # Productivity section
    r += 1
    if prod.get("status") == "OK":
        ws_wf.merge_cells(f"A{r}:D{r}")
        ws_wf.cell(row=r, column=1, value="PRODUCTIVITY METRICS").font = Font(name="Arial", size=12, bold=True, color=DARK_BG)
        r += 1
        prod_items = [
            ("Date Column", prod.get("date_column", "N/A")),
            ("Date Range", prod.get("date_range", "N/A")),
            ("Total Records", prod.get("total_records", 0)),
            ("Daily Average", prod.get("daily_avg", 0)),
            ("Peak Day", f"{prod.get('peak_day', 'N/A')} ({prod.get('peak_count', 0)} items)"),
            ("Monthly Average", prod.get("monthly_avg", 0)),
        ]
        if prod.get("top_performer"):
            prod_items.append(("Top Performer", prod["top_performer"]))

        write_header_row(ws_wf, r, ["Metric", "Value", "", ""])
        r += 1
        for i, (m, v) in enumerate(prod_items):
            write_data_row(ws_wf, r, [m, v, "", ""], stripe=i % 2 == 0)
            r += 1

        # Group breakdown table
        gb = prod.get("group_breakdown", {})
        if gb:
            r += 1
            ws_wf.merge_cells(f"A{r}:D{r}")
            ws_wf.cell(row=r, column=1, value=f"BY {prod.get('group_column', 'GROUP').upper()}").font = Font(name="Arial", size=12, bold=True, color=DARK_BG)
            r += 1
            write_header_row(ws_wf, r, ["Group", "Count", "Daily Avg", "Days Active"])
            r += 1
            for i, (grp, info) in enumerate(gb.items()):
                write_data_row(ws_wf, r, [grp, info["count"], info["daily_avg"], info["days_active"]], stripe=i % 2 == 0)
                r += 1

        # Monthly trend
        monthly = prod.get("monthly_trend", {})
        if monthly:
            r += 1
            ws_wf.merge_cells(f"A{r}:D{r}")
            ws_wf.cell(row=r, column=1, value="MONTHLY TREND").font = Font(name="Arial", size=12, bold=True, color=DARK_BG)
            r += 1
            write_header_row(ws_wf, r, ["Month", "Count", "", ""])
            r += 1
            for i, (month, count) in enumerate(monthly.items()):
                write_data_row(ws_wf, r, [str(month), count, "", ""], stripe=i % 2 == 0)
                r += 1
    else:
        ws_wf.cell(row=r, column=1, value=prod.get("message", "No date column detected")).font = Font(name="Arial", size=11, color=MID_GRAY, italic=True)

    # ── Save ──
    wb.save(output_path)
    return output_path
